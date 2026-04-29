"""Excel report exporter using openpyxl."""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from app.domain.reports.entities import ReportColumn, ReportDefinition


class ExcelReportExporter:
    """Export report data to Excel format with professional formatting."""

    # Style configurations
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    DATA_FONT = Font(name="Calibri", size=10)
    DATA_ALIGNMENT = Alignment(vertical="center")
    
    CURRENCY_FORMAT = '"$"#,##0.00'
    NUMBER_FORMAT = "#,##0"
    PERCENT_FORMAT = "0.00%"
    DATE_FORMAT = "YYYY-MM-DD"
    DATETIME_FORMAT = "YYYY-MM-DD HH:MM:SS"
    
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def export(
        self,
        data: list[dict[str, Any]],
        definition: ReportDefinition,
        title: str | None = None,
    ) -> bytes:
        """
        Export data to Excel format.
        
        Args:
            data: List of dictionaries containing report data
            definition: Report definition with column configuration
            title: Optional report title for the worksheet
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = title or definition.name[:31]  # Excel sheet name limit

        # Determine columns to display
        if definition.columns:
            columns = [c for c in definition.columns if c.visible]
            fields = [c.field for c in columns]
            headers = {c.field: c.label for c in columns}
        else:
            fields = list(data[0].keys()) if data else []
            headers = {f: self._format_header(f) for f in fields}
            columns = []

        if not fields:
            return self._create_empty_workbook()

        # Write title row
        row_num = 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(fields))
        title_cell = ws.cell(row=row_num, column=1)
        title_cell.value = title or definition.name
        title_cell.font = Font(name="Calibri", size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        row_num += 1

        # Write metadata row
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(fields))
        meta_cell = ws.cell(row=row_num, column=1)
        meta_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Rows: {len(data)}"
        meta_cell.font = Font(name="Calibri", size=9, italic=True, color="666666")
        meta_cell.alignment = Alignment(horizontal="center")
        row_num += 2  # Add blank row

        # Write header row
        header_row = row_num
        for col_idx, field in enumerate(fields, start=1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = headers.get(field, field)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        row_num += 1

        # Write data rows
        for row_data in data:
            for col_idx, field in enumerate(fields, start=1):
                cell = ws.cell(row=row_num, column=col_idx)
                value = row_data.get(field)
                
                # Set value and apply formatting
                cell.value = self._convert_value(value)
                cell.font = self.DATA_FONT
                cell.alignment = self.DATA_ALIGNMENT
                cell.border = self.THIN_BORDER
                
                # Apply number format based on column config or value type
                cell.number_format = self._get_number_format(field, value, columns)
            row_num += 1

        # Auto-fit column widths
        self._auto_fit_columns(ws, fields, data, headers)

        # Add autofilter
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(fields))}{row_num - 1}"

        # Freeze header row
        ws.freeze_panes = f"A{header_row + 1}"

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def _format_header(self, field: str) -> str:
        """Convert field name to header label."""
        return field.replace("_", " ").title()

    def _convert_value(self, value: Any) -> Any:
        """Convert value to Excel-compatible type."""
        if value is None:
            return ""
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, (date, datetime)):
            return value
        if isinstance(value, (list, dict)):
            return str(value)
        return value

    def _get_number_format(
        self,
        field: str,
        value: Any,
        columns: list[ReportColumn],
    ) -> str:
        """Determine appropriate number format for a cell."""
        # Check column format specification
        for col in columns:
            if col.field == field and col.format:
                if col.format == "currency":
                    return self.CURRENCY_FORMAT
                elif col.format == "percent":
                    return self.PERCENT_FORMAT
                elif col.format == "number":
                    return self.NUMBER_FORMAT
                elif col.format == "date":
                    return self.DATE_FORMAT
                elif col.format == "datetime":
                    return self.DATETIME_FORMAT

        # Infer from field name
        field_lower = field.lower()
        if any(kw in field_lower for kw in ["amount", "price", "cost", "total", "revenue"]):
            return self.CURRENCY_FORMAT
        if "percent" in field_lower or "rate" in field_lower:
            return self.PERCENT_FORMAT
        if isinstance(value, datetime):
            return self.DATETIME_FORMAT
        if isinstance(value, date):
            return self.DATE_FORMAT
        if isinstance(value, (int, float, Decimal)):
            return self.NUMBER_FORMAT

        return "General"

    def _auto_fit_columns(
        self,
        ws,
        fields: list[str],
        data: list[dict[str, Any]],
        headers: dict[str, str],
    ) -> None:
        """Auto-fit column widths based on content."""
        for col_idx, field in enumerate(fields, start=1):
            # Calculate max width from header and data
            header_width = len(str(headers.get(field, field)))
            
            max_data_width = 0
            for row_data in data[:100]:  # Sample first 100 rows
                value = row_data.get(field, "")
                max_data_width = max(max_data_width, len(str(value)))

            # Set column width with padding
            width = min(max(header_width, max_data_width) + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _create_empty_workbook(self) -> bytes:
        """Create an empty workbook with a message."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "No data available"
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()


class ExcelSummaryExporter(ExcelReportExporter):
    """Export summary reports with charts and aggregations."""

    def export_with_summary(
        self,
        data: list[dict[str, Any]],
        definition: ReportDefinition,
        summary: dict[str, Any],
        title: str | None = None,
    ) -> bytes:
        """
        Export data with summary section at the top.
        
        Args:
            data: Report data rows
            definition: Report definition
            summary: Summary statistics to display
            title: Optional report title
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = title or definition.name[:31]

        row_num = 1

        # Title
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
        title_cell = ws.cell(row=row_num, column=1)
        title_cell.value = title or definition.name
        title_cell.font = Font(name="Calibri", size=16, bold=True)
        row_num += 2

        # Summary section
        ws.cell(row=row_num, column=1).value = "Summary"
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
        row_num += 1

        for key, value in summary.items():
            label_cell = ws.cell(row=row_num, column=1)
            label_cell.value = self._format_header(key)
            label_cell.font = Font(bold=True)
            
            value_cell = ws.cell(row=row_num, column=2)
            value_cell.value = self._convert_value(value)
            if isinstance(value, (int, float, Decimal)):
                if "amount" in key.lower() or "revenue" in key.lower():
                    value_cell.number_format = self.CURRENCY_FORMAT
                else:
                    value_cell.number_format = self.NUMBER_FORMAT
            row_num += 1

        row_num += 1  # Blank row

        # Data section header
        ws.cell(row=row_num, column=1).value = "Detail Data"
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
        row_num += 1

        # Determine columns
        if definition.columns:
            columns = [c for c in definition.columns if c.visible]
            fields = [c.field for c in columns]
            headers = {c.field: c.label for c in columns}
        else:
            fields = list(data[0].keys()) if data else []
            headers = {f: self._format_header(f) for f in fields}
            columns = []

        if not fields:
            return self._create_empty_workbook()

        # Header row
        header_row = row_num
        for col_idx, field in enumerate(fields, start=1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = headers.get(field, field)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        row_num += 1

        # Data rows
        for row_data in data:
            for col_idx, field in enumerate(fields, start=1):
                cell = ws.cell(row=row_num, column=col_idx)
                value = row_data.get(field)
                cell.value = self._convert_value(value)
                cell.font = self.DATA_FONT
                cell.border = self.THIN_BORDER
                cell.number_format = self._get_number_format(field, value, columns)
            row_num += 1

        # Auto-fit columns
        self._auto_fit_columns(ws, fields, data, headers)

        # Add autofilter
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(fields))}{row_num - 1}"

        # Save
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
